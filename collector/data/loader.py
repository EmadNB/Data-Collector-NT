"""Data loading functions for ENTSO-E PEMMDB, PECD, and network files."""

from __future__ import annotations

import re
import openpyxl
from openpyxl.utils import get_column_letter
from functools import lru_cache
import numpy as np
import pandas as pd

# Other Non-RES has up to 27 type columns (C..AC) on its PEMMDB sheet.
_OTHER_NONRES_COLS = [get_column_letter(3 + i) for i in range(27)]


@lru_cache(maxsize=16)
def _excel(path: str) -> pd.ExcelFile:
    """Return a cached ``pd.ExcelFile`` for *path*.

    Reading individual cells with ``pd.read_excel(path, ...)`` re-opens and
    re-parses the whole workbook every call (~65 ms each). Passing a shared
    ``ExcelFile`` instead unzips/parses the workbook once, making per-cell reads
    ~7x cheaper. Exceptions (e.g. missing file) are not cached by ``lru_cache``.
    """
    return pd.ExcelFile(path)


def clear_excel_cache() -> None:
    """Drop cached workbook handles so a fresh run picks up any input changes."""
    _excel.cache_clear()
    _dsr_col_count.cache_clear()

from collector.utils.config import (
    FILEPATH_CO2_FACTORS,
    FILEPATH_COMMON_DATA,
    FILEPATH_COMMODITY_PRICES,
    FILEPATH_NETWORKS,
    FILEPATH_STORAGES,
    FILEPATH_TERMINALS,
    GAS_UNIT_FACTOR,
    HYDRO_FILE_TEMPLATES,
    HYDRO_SCALE_FACTOR,
    HYDRO_SHEET_NAMES,
    HYDRO_TARGET_LENGTHS,
    PECD_FILE_TEMPLATES,
    RESERVE_COLUMNS,
    SOLAR_ROOFTOP_TARGET_LEN,
    TECH_CHAR_COLUMNS,
    TECH_COLUMNS,
    DSR_DEFAULT_COUNT,
    build_tech_columns,
)
from collector.utils.helpers import get_co2_usecols, get_pemmdb_filepath


@lru_cache(maxsize=256)
def _dsr_col_count(filepath: str) -> int:
    """Number of DSR type columns (C..) on a zone's DSR sheet, 0 if unavailable.

    Uses a fresh read-only workbook: reading ``max_column`` from the shared
    cached ``_excel`` handle is unreliable once its worksheet has been streamed.
    The integer result is cached so detection happens once per file.
    """
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        try:
            n = wb["DSR"].max_column
        finally:
            wb.close()
        return max((n or 2) - 2, 0)
    except Exception:
        return 0


def _dsr_count_for(selected_zones: list[str], scenario: int) -> int:
    """Max DSR type-column count across *selected_zones* (>= the default 10)."""
    counts = [_dsr_col_count(get_pemmdb_filepath(z, scenario)) for z in selected_zones]
    return max(counts + [DSR_DEFAULT_COUNT])


def _dsr_cols(n_dsr: int) -> list[str]:
    """Excel column letters C.. for *n_dsr* DSR type columns."""
    return [get_column_letter(3 + i) for i in range(n_dsr)]


# ---------------------------------------------------------------------------
# Node / network raw loaders
# ---------------------------------------------------------------------------


def load_nodes(filepath: str = FILEPATH_NETWORKS, sheet_name: str = "Nodes") -> pd.DataFrame:
    """Load the node (zone) reference table from an Excel workbook.

    Args:
        filepath (str): Path to the Excel file containing the Nodes sheet.
            Defaults to ``Data/Networks.xlsx``.
        sheet_name (str): Name of the worksheet to read. Defaults to
            ``'Nodes'``.

    Returns:
        pd.DataFrame: DataFrame with at least the columns ``Code``,
            ``Latitude``, and ``Longitude``.

    Raises:
        FileNotFoundError: When *filepath* does not exist.

    Example:
        >>> nodes = load_nodes()
        >>> list(nodes.columns)[:3]
        ['Code', 'Latitude', 'Longitude']
    """
    return pd.read_excel(filepath, sheet_name=sheet_name)


def load_network_edges(
    scenario: int,
    filepath: str = FILEPATH_NETWORKS,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Load electricity, gas, and hydrogen network edge tables from Excel.

    Args:
        scenario (int): Scenario year (``2030``, ``2040``, or ``2050``).
        filepath (str): Path to ``Networks.xlsx``.

    Returns:
        tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]: Three DataFrames for
            electricity lines, gas pipelines, and hydrogen pipelines
            respectively.

    Raises:
        FileNotFoundError: When *filepath* does not exist.

    Example:
        >>> edges_e, edges_g, edges_h = load_network_edges(2030)
    """
    # Only one electricity network exists, so its sheet is period-independent
    # (named simply "Lines_E"); gas and hydrogen remain per-scenario.
    edges_e = pd.read_excel(filepath, sheet_name="Lines_E")
    edges_g = pd.read_excel(filepath, sheet_name=f"Lines_G ({scenario})")
    edges_h = pd.read_excel(filepath, sheet_name=f"Lines_H ({scenario})")
    return edges_e, edges_g, edges_h


def load_network_storages(
    scenario: int,
    filepath: str = FILEPATH_STORAGES,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Load gas and hydrogen storage capacity tables from Excel.

    Args:
        scenario (int): Scenario year.
        filepath (str): Path to ``Storages.xlsx``.

    Returns:
        tuple[pd.DataFrame, pd.DataFrame]: Gas storages DataFrame and
            hydrogen storages DataFrame.

    Raises:
        FileNotFoundError: When *filepath* does not exist.

    Example:
        >>> storages_g, storages_h = load_network_storages(2030)
    """
    storages_g = pd.read_excel(filepath, sheet_name=f"Storage_G ({scenario})")
    storages_h = pd.read_excel(filepath, sheet_name=f"Storage_H ({scenario})")
    return storages_g, storages_h


def load_network_terminals(
    scenario: int,
    filepath: str = FILEPATH_TERMINALS,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Load gas and hydrogen terminal capacity tables from Excel.

    Args:
        scenario (int): Scenario year.
        filepath (str): Path to ``Terminals.xlsx``.

    Returns:
        tuple[pd.DataFrame, pd.DataFrame]: Gas terminals DataFrame and
            hydrogen terminals DataFrame.

    Raises:
        FileNotFoundError: When *filepath* does not exist.

    Example:
        >>> terminals_g, terminals_h = load_network_terminals(2030)
    """
    terminals_g = pd.read_excel(filepath, sheet_name=f"Terminal_G ({scenario})")
    terminals_h = pd.read_excel(filepath, sheet_name=f"Terminal_H ({scenario})")
    return terminals_g, terminals_h


# ---------------------------------------------------------------------------
# PEMMDB loaders
# ---------------------------------------------------------------------------


def load_tech_capacities(
    node_df: pd.DataFrame,
    selected_zones: list[str],
    scenario: int,
    selected_hours: int,
) -> pd.DataFrame:
    """Load installed technology capacity data from PEMMDB Excel files.

    Reads per-technology MW / MWh values and selected hourly time series
    (exports, DSR, other RES) for every zone in *selected_zones* from the
    corresponding PEMMDB workbook.  Missing files are handled gracefully by
    filling zeros for all columns.

    Args:
        node_df (pd.DataFrame): Nodes table whose ``Code`` column defines
            the universe of available zone codes.
        selected_zones (list[str]): Subset of zone codes to collect data for.
        scenario (int): Scenario year (``2030``, ``2040``, or ``2050``).
        selected_hours (int): Number of hourly values to read for time-series
            columns (e.g. ``8736``).

    Returns:
        pd.DataFrame: One row per zone with columns defined by
            :data:`~collector.utils.config.TECH_COLUMNS`.

    Example:
        >>> nodes = load_nodes()
        >>> cap_df = load_tech_capacities(nodes, ["ES00", "PT00"], 2030, 8736)
        >>> "Nuclear (MW)" in cap_df.columns
        True
    """
    n_dsr = _dsr_count_for(selected_zones, scenario)
    columns = build_tech_columns(n_dsr)
    tech_rows: list[dict] = []
    for code in node_df["Code"]:
        if code not in selected_zones:
            continue
        filepath = get_pemmdb_filepath(code, scenario)
        data: dict = {"Code": code}
        try:
            _read_thermal_capacities(filepath, data)
            _read_hydro_capacities(filepath, data)
            _read_res_capacities(filepath, data, n_dsr)
            _read_storage_capacities(filepath, data)
            _read_timeseries_capacities(filepath, data, selected_hours, n_dsr)
            tech_rows.append(data)
            print(f"Technology capacities for {code}: OK")
        except FileNotFoundError:
            print(f"Technology capacities for {code}: file not found – zeros used")
            data.update({col: 0 for col in columns[1:]})
            tech_rows.append(data)
    return pd.DataFrame(tech_rows, columns=columns)


def _read_scalar(filepath: str, sheet: str, col: str, row: int,
                 default: object = 0.0) -> object:
    """Read a single cell (sheet/col/row) from an Excel file, safely.

    Returns *default* when the sheet is missing, the column is out of bounds, or
    the requested row lies beyond the sheet's extent — situations that occur for
    some countries whose PEMMDB sheets are shorter or narrower than expected.
    Without this, ``read_excel`` raises ``ParserError`` (out-of-bounds usecols)
    or ``.iat[0, 0]`` raises ``IndexError`` on an empty frame.
    """
    try:
        df = pd.read_excel(
            _excel(filepath), sheet_name=sheet, usecols=col, header=None, skiprows=row, nrows=1
        )
        if df is None or df.shape[1] == 0 or len(df) == 0:
            return default
        return df.iat[0, 0]
    except Exception:
        return default


def _read_thermal_capacities(filepath: str, data: dict) -> None:
    """Populate *data* with all thermal technology capacity values."""
    def _cell(sheet: str, col: str, row: int) -> float:
        return _read_scalar(filepath, sheet, col, row)

    data["Nuclear (MW)"]           = _cell("Thermal", "C", 11)
    data["Hard Coal (old1) (MW)"]  = _cell("Thermal", "C", 13)
    data["Hard Coal (old2) (MW)"]  = _cell("Thermal", "C", 15)
    data["Hard Coal (new) (MW)"]   = _cell("Thermal", "C", 17)
    data["Hard Coal (ccs) (MW)"]   = _cell("Thermal", "C", 19)
    data["Lignite (old1) (MW)"]    = _cell("Thermal", "C", 21)
    data["Lignite (old2) (MW)"]    = _cell("Thermal", "C", 23)
    data["Lignite (new) (MW)"]     = _cell("Thermal", "C", 25)
    data["Lignite (ccs) (MW)"]     = _cell("Thermal", "C", 27)
    data["Gas (conv_old1) (MW)"]   = _cell("Thermal", "C", 29)
    data["Gas (conv_old2) (MW)"]   = _cell("Thermal", "C", 31)
    data["Gas (ccgt_old1) (MW)"]   = _cell("Thermal", "C", 33)
    data["Gas (ccgt_old2) (MW)"]   = _cell("Thermal", "C", 35)
    data["Gas (ccgt_new) (MW)"]    = _cell("Thermal", "C", 37)
    data["Gas (ccgt_ccs) (MW)"]    = _cell("Thermal", "C", 39)
    data["Gas (ocgt_old) (MW)"]    = _cell("Thermal", "C", 41)
    data["Gas (ocgt_new) (MW)"]    = _cell("Thermal", "C", 43)
    data["Light Oil (MW)"]         = _cell("Thermal", "C", 45)
    data["Heavy oil (old1) (MW)"]  = _cell("Thermal", "C", 47)
    data["Heavy oil (old2) (MW)"]  = _cell("Thermal", "C", 49)
    data["Oil shale (old) (MW)"]   = _cell("Thermal", "C", 51)
    data["Oil shale (new) (MW)"]   = _cell("Thermal", "C", 53)
    data["Gas (ccgt_pre1) (MW)"]   = _cell("Thermal", "C", 55)
    data["Gas (ccgt_pre2) (MW)"]   = _cell("Thermal", "C", 57)
    data["Hydrogen (fc) (MW)"]     = _cell("Thermal", "C", 59)
    data["Hydrogen (ccgt) (MW)"]   = _cell("Thermal", "C", 61)


def _read_hydro_capacities(filepath: str, data: dict) -> None:
    """Populate *data* with hydro technology capacity values."""
    def _cell(col: str, row: int, scale: float = 1.0) -> float:
        return _read_scalar(filepath, "Hydro", col, row) * scale

    data["Hydro (river) (MW)"]            = _cell("B", 8)
    data["Hydro (pondage) (MWh)"]         = _cell("B", 10, 1000)
    data["Hydro (pondage) (MW)"]          = _cell("B", 11)
    data["Hydro (reservoir) (MWh)"]       = _cell("B", 13, 1000)
    data["Hydro (reservoir) (MW)"]        = _cell("B", 14)
    data["Hydro (open_ps) (MWh)"]         = _cell("B", 16, 1000)
    data["Hydro (open_ps_turbine) (MW)"]  = _cell("B", 17)
    data["Hydro (open_ps_pump) (MW)"]     = _cell("B", 18)
    data["Hydro (closed_ps) (MWh)"]       = _cell("B", 20, 1000)
    data["Hydro (closed_ps_turbine) (MW)"] = _cell("B", 21)
    data["Hydro (closed_ps_pump) (MW)"]   = _cell("B", 22)


def _read_res_capacities(filepath: str, data: dict, n_dsr: int = DSR_DEFAULT_COUNT) -> None:
    """Populate *data* with RES and additional technology capacity values."""
    def _cell(sheet: str, col: str, row: int, scale: float = 1.0) -> float:
        return _read_scalar(filepath, sheet, col, row) * scale

    data["Wind (onshore) (MW)"]                  = _cell("Wind",      "B", 7, 1000)
    data["Wind (offshore) (MW)"]                 = _cell("Wind",      "B", 8, 1000)
    data["Solar (thermal) (MW)"]                 = _cell("Solar",     "B", 7, 1000)
    data["Solar (MW)"]                           = _cell("Solar",     "B", 8, 1000)
    data["Solar (rooftop) (MW)"]                 = _cell("Solar",     "B", 9, 1000)
    data["Solar (thermal_with_storage) (MW)"]    = _cell("Solar",     "B", 10, 1000)
    data["Solar (thermal_with_storage) (MWh)"]   = _cell("Solar",     "B", 11, 1000)
    for _i, _col in enumerate(_OTHER_NONRES_COLS):
        data[f"Other Non-RES{_i+1} (MW)"] = _cell("Other Non-RES", _col, 8)
    data["Other RES (biomass) (MW)"]             = _cell("Other RES", "E", 8)
    data["Other RES (geothermal) (MW)"]          = _cell("Other RES", "F", 8)
    data["Other RES (marine) (MW)"]              = _cell("Other RES", "G", 8)
    data["Other RES (waste) (MW)"]               = _cell("Other RES", "H", 8)
    data["Other RES (unknown) (MW)"]             = _cell("Other RES", "I", 8)
    for _i, _col in enumerate(_dsr_cols(n_dsr)):
        data[f"DSR{_i+1} (MW)"] = _cell("DSR", _col, 8)


def _read_storage_capacities(filepath: str, data: dict) -> None:
    """Populate *data* with battery and electrolyser capacity values."""
    def _cell(sheet: str, col: str, row: int) -> float:
        return _read_scalar(filepath, sheet, col, row)

    data["Battery (MWh)"]      = _cell("Battery",     "E", 11)
    data["Electrolyser (MW)"]  = _cell("Electrolyser", "C", 11)
    data["Electrolyser (MWh)"] = _cell("Electrolyser", "F", 11)


def _read_timeseries_capacities(filepath: str, data: dict, selected_hours: int,
                                n_dsr: int = DSR_DEFAULT_COUNT) -> None:
    """Populate *data* with hourly time-series export / profile columns.

    Reads all of a sheet's needed columns in a single call. Per-column reads on a
    read-only workbook re-iterate the whole sheet each time (~1.7 s per column),
    so batching a sheet's block of columns into one read is ~10x faster.
    """
    _zeros = lambda: np.zeros((selected_hours, 1))

    def _block(sheet: str, first: str, last: str, row: int) -> pd.DataFrame:
        # One read for the whole column range; empty frame on any failure.
        try:
            return pd.read_excel(
                _excel(filepath), sheet_name=sheet, usecols=f"{first}:{last}",
                header=None, skiprows=row, nrows=selected_hours,
            )
        except Exception:
            return pd.DataFrame()

    def _one(sheet: str, col: str, row: int) -> np.ndarray:
        try:
            return pd.read_excel(
                _excel(filepath), sheet_name=sheet, usecols=col, header=None,
                skiprows=row, nrows=selected_hours,
            ).to_numpy()
        except Exception:
            return _zeros()

    def _assign(df: pd.DataFrame, i: int) -> np.ndarray:
        return df.iloc[:, i:i + 1].to_numpy() if i < df.shape[1] else _zeros()

    # Other Non-RES: 27 type columns C..AC in one read.
    onr = _block("Other Non-RES", "C", _OTHER_NONRES_COLS[-1], 18)
    for _i in range(len(_OTHER_NONRES_COLS)):
        data[f"Other Non-RES{_i+1} (MW/h)"] = _assign(onr, _i)

    data["Exports_non_ENTSOe (MW/h)"] = -_one("Exchanges", "C", 28)

    # DSR: read only this zone's actual width (avoids costly out-of-range reads);
    # pad any remaining DSR{i} columns up to n_dsr with zeros.
    _w = _dsr_col_count(filepath)
    dsr = _block("DSR", "C", get_column_letter(2 + _w), 15) if _w > 0 else pd.DataFrame()
    for _i in range(n_dsr):
        data[f"DSR{_i+1} (MW/h)"] = _assign(dsr, _i)

    # Other RES: biomass/geothermal/marine/waste/unknown = columns E..I in one read.
    ores = _block("Other RES", "E", "I", 10)
    for _i, _nm in enumerate(("biomass", "geothermal", "marine", "waste", "unknown")):
        data[f"Other RES ({_nm}) (MW/h)"] = _assign(ores, _i)


def load_tech_characteristics(
    node_df: pd.DataFrame,
    selected_zones: list[str],
    scenario: int,
) -> pd.DataFrame:
    """Load technology characteristic data (outage rates, ramp rates, etc.).

    Reads per-unit characteristics for all thermal technologies, as well as
    DSR, Battery, and Electrolyser units, from the PEMMDB workbooks.  CO2
    factors and fuel prices are drawn from the common reference files.

    Args:
        node_df (pd.DataFrame): Nodes table (used for zone iteration).
        selected_zones (list[str]): Zone codes to collect data for.
        scenario (int): Scenario year.

    Returns:
        pd.DataFrame: One row per zone with columns defined by
            :data:`~collector.utils.config.TECH_CHAR_COLUMNS`.

    Example:
        >>> nodes = load_nodes()
        >>> char_df = load_tech_characteristics(nodes, ["ES00"], 2030)
        >>> "Ramp-Up Rate (MW/h)" in char_df.columns
        True
    """
    co2_col = get_co2_usecols(scenario)
    n_dsr = _dsr_count_for(selected_zones, scenario)
    tech_char_rows: list[dict] = []

    for code in node_df["Code"]:
        if code not in selected_zones:
            continue
        filepath = get_pemmdb_filepath(code, scenario)
        try:
            data_char = _read_single_zone_characteristics(filepath, co2_col, code, n_dsr)
            tech_char_rows.append(data_char)
            print(f"Technology characteristics for {code}: OK")
        except FileNotFoundError:
            print(f"Technology characteristics for {code}: file not found – skipped")

    return pd.DataFrame(tech_char_rows, columns=TECH_CHAR_COLUMNS)


def _read_single_zone_characteristics(
    filepath: str, co2_col: str, code: str, n_dsr: int = DSR_DEFAULT_COUNT
) -> dict:
    """Read all technology characteristic fields for one zone."""
    def _arr(col: str, row: int, nrows: int = 52) -> np.ndarray:
        try:
            return pd.read_excel(
                _excel(filepath), sheet_name="Thermal", usecols=col, header=None,
                skiprows=row, nrows=nrows,
            ).to_numpy()
        except Exception:
            return np.empty((0, 1))

    def _cell(filepath_: str, sheet: str, col: str, row: int) -> object:
        return _read_scalar(filepath_, sheet, col, row)

    dc: dict = {"Code": code}

    raw = _arr("D", 11); dc["Number of Units"]                  = raw[~np.isnan(raw.astype(float))]
    raw = _arr("E", 11); dc["Number of Biofuel Units"]          = raw[~np.isnan(raw.astype(float))]
    raw = _arr("F", 11); dc["Biofuel Usage (%)"]                = raw[~np.isnan(raw.astype(float))]
    raw = _arr("H:S", 11)
    dc["Must Run (Number of units)"] = raw[::2]
    dc["Must Run (%)"]               = raw[1::2]
    raw = _arr("AG", 11); dc["Annual Forced Outage (%)"]        = raw[~np.isnan(raw.astype(float))]
    raw = _arr("AH", 11); dc["Annual Forced Outage (Days)"]     = raw[~np.isnan(raw.astype(float))]
    raw = _arr("AI", 11); dc["Annual Forced Outage in Winter (%)"] = raw[~np.isnan(raw.astype(float))]
    raw = _arr("AJ", 11); dc["Minimum Stable Power (%)"]        = raw[~np.isnan(raw.astype(float))]
    raw = _arr("AK", 11); dc["Ramp-Up Rate (MW/h)"]             = raw[~np.isnan(raw.astype(float))]
    raw = _arr("AL", 11); dc["Ramp-Down Rate (MW/h)"]           = raw[~np.isnan(raw.astype(float))]
    raw = _arr("AM", 11); dc["Fixed Generation Reduction (%)"]  = raw[~np.isnan(raw.astype(float))]
    raw = _arr("AP", 11); dc["Maximum Number of Units in Maintenace"] = raw[~np.isnan(raw.astype(float))]

    dc["CO2 Factor (ton/MWh)"] = (
        pd.read_excel(_excel(FILEPATH_CO2_FACTORS), sheet_name="CO2 emission factor",
                      usecols=co2_col, header=None, skiprows=4, nrows=26).to_numpy() * 0.0036
    )
    dc["Efficiency (%)"] = pd.read_excel(
        _excel(FILEPATH_COMMON_DATA), sheet_name="Common Data", usecols="F",
        header=None, skiprows=14, nrows=26,
    ).to_numpy()
    dc["Price (EUR/MWh)"] = pd.read_excel(
        _excel(FILEPATH_COMMON_DATA), sheet_name="Common Data", usecols="H",
        header=None, skiprows=14, nrows=26,
    ).to_numpy()

    zeros26 = np.zeros(26)
    dc["Net maximum capacity - generation perspective (MW)"] = zeros26.copy()
    dc["Net maximum capacity - demand perspective (MW)"]     = zeros26.copy()

    # Other Non-RES1..27 (each successive column C..AC in the Other Non-RES sheet)
    for _col in _OTHER_NONRES_COLS:
        for key in ("Fixed Generation Reduction (%)", "Ramp-Up Rate (MW/h)", "Ramp-Down Rate (MW/h)"):
            dc[key] = np.append(dc[key], 0)
        dc["Number of Units"]   = np.append(dc["Number of Units"],   _cell(filepath, "Other Non-RES", _col, 9))
        dc["Price (EUR/MWh)"]   = np.append(dc["Price (EUR/MWh)"],   _cell(filepath, "Other Non-RES", _col, 12))
        dc["Efficiency (%)"]    = np.append(dc["Efficiency (%)"],    _cell(filepath, "Other Non-RES", _col, 13))
        dc["CO2 Factor (ton/MWh)"] = np.append(dc["CO2 Factor (ton/MWh)"], _cell(filepath, "Other Non-RES", _col, 14))
        dc["Net maximum capacity - generation perspective (MW)"] = np.append(dc["Net maximum capacity - generation perspective (MW)"], 0)
        dc["Net maximum capacity - demand perspective (MW)"]     = np.append(dc["Net maximum capacity - demand perspective (MW)"], 0)

    # DSR1..n_dsr (each successive column in the DSR sheet)
    for _col in _dsr_cols(n_dsr):
        for key in ("Fixed Generation Reduction (%)", "Ramp-Up Rate (MW/h)", "Ramp-Down Rate (MW/h)"):
            dc[key] = np.append(dc[key], 0)
        dc["Number of Units"] = np.append(dc["Number of Units"], _cell(filepath, "DSR", _col, 9))
        dc["Price (EUR/MWh)"] = np.append(dc["Price (EUR/MWh)"], _cell(filepath, "DSR", _col, 11))
        dc["Efficiency (%)"]  = np.append(dc["Efficiency (%)"], 0)
        dc["CO2 Factor (ton/MWh)"] = np.append(dc["CO2 Factor (ton/MWh)"], 0)
        dc["Net maximum capacity - generation perspective (MW)"] = np.append(dc["Net maximum capacity - generation perspective (MW)"], 0)
        dc["Net maximum capacity - demand perspective (MW)"]     = np.append(dc["Net maximum capacity - demand perspective (MW)"], 0)

    # Battery
    dc["Fixed Generation Reduction (%)"] = np.append(dc["Fixed Generation Reduction (%)"], 0)
    dc["Ramp-Up Rate (MW/h)"]   = np.append(dc["Ramp-Up Rate (MW/h)"],   _cell(filepath, "Battery", "H", 11))
    dc["Ramp-Down Rate (MW/h)"] = np.append(dc["Ramp-Down Rate (MW/h)"], _cell(filepath, "Battery", "I", 11))
    dc["Number of Units"] = np.append(dc["Number of Units"], _cell(filepath, "Battery", "F", 11))
    dc["Price (EUR/MWh)"] = np.append(dc["Price (EUR/MWh)"], 0)
    dc["Efficiency (%)"]  = np.append(dc["Efficiency (%)"],  _cell(filepath, "Battery", "G", 11))
    dc["CO2 Factor (ton/MWh)"] = np.append(dc["CO2 Factor (ton/MWh)"], 0)
    dc["Net maximum capacity - generation perspective (MW)"] = np.append(dc["Net maximum capacity - generation perspective (MW)"], _cell(filepath, "Battery", "C", 11))
    dc["Net maximum capacity - demand perspective (MW)"]     = np.append(dc["Net maximum capacity - demand perspective (MW)"], _cell(filepath, "Battery", "D", 11))

    # Electrolyser
    dc["Fixed Generation Reduction (%)"] = np.append(dc["Fixed Generation Reduction (%)"], _cell(filepath, "Electrolyser", "I", 11))
    dc["Ramp-Up Rate (MW/h)"]   = np.append(dc["Ramp-Up Rate (MW/h)"],   _cell(filepath, "Electrolyser", "G", 11))
    dc["Ramp-Down Rate (MW/h)"] = np.append(dc["Ramp-Down Rate (MW/h)"], _cell(filepath, "Electrolyser", "H", 11))
    dc["Number of Units"] = np.append(dc["Number of Units"], _cell(filepath, "Electrolyser", "D", 11))
    dc["Price (EUR/MWh)"] = np.append(dc["Price (EUR/MWh)"], 0)
    dc["Efficiency (%)"]  = np.append(dc["Efficiency (%)"],  _cell(filepath, "Electrolyser", "E", 11))
    dc["CO2 Factor (ton/MWh)"] = np.append(dc["CO2 Factor (ton/MWh)"], 0)
    dc["Net maximum capacity - generation perspective (MW)"] = np.append(dc["Net maximum capacity - generation perspective (MW)"], 0)
    dc["Net maximum capacity - demand perspective (MW)"]     = np.append(dc["Net maximum capacity - demand perspective (MW)"], 0)

    return dc


def load_reserve_requirements(
    node_df: pd.DataFrame,
    selected_zones: list[str],
    scenario: int,
) -> pd.DataFrame:
    """Load FCR and FRR reserve requirement data from PEMMDB workbooks.

    Args:
        node_df (pd.DataFrame): Nodes table used for zone iteration.
        selected_zones (list[str]): Zone codes to collect.
        scenario (int): Scenario year.

    Returns:
        pd.DataFrame: One row per zone with columns defined by
            :data:`~collector.utils.config.RESERVE_COLUMNS`.

    Example:
        >>> nodes = load_nodes()
        >>> res_df = load_reserve_requirements(nodes, ["ES00"], 2030)
        >>> "Total (FCR) (MW/h)" in res_df.columns
        True
    """
    rows: list[dict] = []
    for code in node_df["Code"]:
        if code not in selected_zones:
            continue
        filepath = get_pemmdb_filepath(code, scenario)
        try:
            def _cell(row: int) -> float:
                return _read_scalar(filepath, "Reserves", "C", row)

            # Reserves sheet orders each block Total -> Thermal -> Hydro:
            #   FCR: Total=C10, Thermal=C11, Hydro=C12
            #   FRR: Total=C16, Thermal=C17, Hydro=C18
            # (_cell(n) reads C{n+1}).
            rows.append({
                "Code":                  code,
                "Total (FCR) (MW/h)":    _cell(9),
                "Thermal (FCR) (MW/h)":  _cell(10),
                "Hydro (FCR) (MW/h)":    _cell(11),
                "Total (FRR) (MW/h)":    _cell(15),
                "Thermal (FRR) (MW/h)":  _cell(16),
                "Hydro (FRR) (MW/h)":    _cell(17),
            })
            print(f"Reserve requirements for {code}: OK")
        except FileNotFoundError:
            print(f"Reserve requirements for {code}: file not found – skipped")

    return pd.DataFrame(rows, columns=RESERVE_COLUMNS)


# ---------------------------------------------------------------------------
# Cross-border exchange results loader
# ---------------------------------------------------------------------------


def load_crossborder_exchanges(
    scenario: int,
    selected_hours: int,
    filtered_edges_e_df: pd.DataFrame,
    selected_zones: list[str],
) -> pd.DataFrame:
    """Load cross-border electricity exchange flows from the MM output file.

    Reads the ``Crossborder exchanges`` worksheet via openpyxl (random-access)
    for any exchange columns that involve a zone in *selected_zones* and a
    neighbouring zone that appears in the filtered electricity edge table.
    Flows *from* a selected zone are positive; flows *into* a selected zone
    are negated.

    Args:
        scenario (int): Scenario year used to build the file path.
        selected_hours (int): Number of hourly rows to read.
        filtered_edges_e_df (pd.DataFrame): Electricity edge table already
            filtered to the study perimeter; must contain ``Start_Node`` and
            ``End_Node`` columns.
        selected_zones (list[str]): Zone codes that define the study area.

    Returns:
        pd.DataFrame: Columns named ``Exports_<zone>_<neighbour> (MW/h)`` with
            one row per hour. External source nodes (codes starting with "X",
            e.g. XRU00, XSA00, XTN00, XMD00, XBACE — not in the edge table) are
            summed into the zone's single ``Exports_<zone>_XX (MW/h)`` column.

    Raises:
        FileNotFoundError: When the MM output workbook is not found.

    Example:
        >>> df = load_crossborder_exchanges(2030, 8736, edges_df, ["ES00"])
    """
    filepath = f"inputs/MMStandardOutputFile_NT{scenario}_Plexos_CY2009_2.5_v40.xlsx"
    unique_nodes = pd.unique(
        pd.concat([
            filtered_edges_e_df["Start_Node"],
            filtered_edges_e_df["End_Node"],
        ]).astype(str)
    )
    unique_nodes_no_selected = [n for n in unique_nodes if n not in [str(z) for z in selected_zones]]

    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb["Crossborder exchanges"]

    header_row = 11
    headers: list[tuple[int, str]] = []
    col_idx = 3
    while True:
        val = ws.cell(row=header_row, column=col_idx).value
        if val is None or str(val).strip() == "":
            break
        headers.append((col_idx, str(val).strip()))
        col_idx += 1

    # (source col, output column name, direction). A normal neighbour present in
    # the network edge table becomes "Exports_<zone>_<node>"; every external
    # source node (codes starting with "X" — e.g. XRU00, XSA00, XTN00, XMD00,
    # XBACE — not in the edge table) is summed into the zone's single
    # "Exports_<zone>" column. Sign: "from" (zone is start) positive, "to" (zone
    # is end) negated.
    specs: list[tuple[int, str, str]] = []
    for col, header in headers:
        for zone in selected_zones:
            zone_str = str(zone)
            if header.startswith(f"{zone_str}->"):
                node = header.split("->", 1)[1]
                direction = "from"
            elif header.endswith(f"->{zone_str}"):
                node = header.split("->", 1)[0]
                direction = "to"
            else:
                continue
            if str(node).startswith("X"):
                name = f"Exports_{zone_str}_XX (MW/h)"        # aggregated external sources
            elif node in unique_nodes_no_selected:
                name = f"Exports_{zone_str}_{node} (MW/h)"
            else:
                continue
            specs.append((col, name, direction))

    row_start = header_row + 1
    row_end = row_start + selected_hours
    col_data: dict[int, list] = {}
    for col in sorted({c for c, _, _ in specs}):
        col_data[col] = [
            v[0] for v in ws.iter_rows(
                min_row=row_start, max_row=row_end - 1,
                min_col=col, max_col=col, values_only=True,
            )
        ]

    wb.close()

    export_df_dict: dict[str, list] = {}
    for col, name, direction in specs:
        raw = col_data[col]
        values = raw if direction == "from" else [-v if v is not None else None for v in raw]
        if name in export_df_dict:  # sum external sources aggregated on one zone
            export_df_dict[name] = [(a or 0) + (b or 0) for a, b in zip(export_df_dict[name], values)]
        else:
            export_df_dict[name] = values

    return pd.DataFrame(export_df_dict)


def load_crossborder_h2_exchanges(
    scenario: int,
    selected_hours: int,
    selected_zones: list[str],
    main_zone_map: dict[str, str],
) -> pd.DataFrame:
    """Load cross-border hydrogen exchange flows from the MM output file.

    Reads the ``Crossborder H2 exchanges`` worksheet, whose headers are
    country-level H2 nodes (e.g. ``AT_H2->DE_H2``). For every flow between a
    selected country and a non-selected neighbour, a column is produced named
    ``H2Exports_<main zone>_<neighbour> (MW/h)``, where *<main zone>* is the
    selected country's main H2 node (from *main_zone_map*, as used for H2
    demand). A country neighbour appears as ``<CC>00``; all external source nodes
    (``XDZ``, ``XMA``, ``XNO``, ``XUA``, ``XAmmonia`` …) are summed into the main
    zone's single ``H2Exports_<main zone>_XX (MW/h)`` column. Sign follows the export
    convention — a flow *from* the selected country is positive, a flow *into* it
    is negated — mirroring :func:`load_crossborder_exchanges`. ``IB*``
    interconnector hubs are resolved end to end (``A->IBIT->IT`` is treated as
    ``A->IT``).

    Args:
        scenario (int): Scenario year used to build the file path.
        selected_hours (int): Number of hourly rows to read.
        selected_zones (list[str]): Zone codes that define the study area.
        main_zone_map (dict[str, str]): Country prefix -> main H2 zone code.

    Returns:
        pd.DataFrame: One column per selected-country/neighbour H2 flow, one row
            per hour. Empty if the worksheet is absent.

    Raises:
        FileNotFoundError: When the MM output workbook is not found.
    """
    filepath = f"inputs/MMStandardOutputFile_NT{scenario}_Plexos_CY2009_2.5_v40.xlsx"
    selected_countries = {str(z)[:2] for z in selected_zones}

    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    if "Crossborder H2 exchanges" not in wb.sheetnames:
        wb.close()
        return pd.DataFrame()
    ws = wb["Crossborder H2 exchanges"]

    header_row = 11
    headers: list[tuple[int, str]] = []
    col_idx = 3
    while True:
        val = ws.cell(row=header_row, column=col_idx).value
        if val is None or str(val).strip() == "":
            break
        headers.append((col_idx, str(val).strip()))
        col_idx += 1

    def _cc(node: str) -> str:
        n = str(node).strip()
        return n[:-3] if n.endswith("_H2") else n

    # Resolve IB* interconnector hubs (e.g. IBIT_H2, IBFI_H2), which pass H2
    # between a source side (A->hub) and a sink side (hub->B). Each "A->hub"
    # segment becomes a direct "A->B" edge carrying the A->hub value, for every
    # sink B; the aggregate "hub->B" edges are dropped (the sources carry the
    # flow). Example: AT_H2->IBIT_H2 plus IBIT_H2->IT_H2 -> AT_H2->IT_H2.
    def _is_hub(node: str) -> bool:
        return str(node).startswith("IB") and str(node).endswith("_H2")

    hub_sinks: dict[str, list[str]] = {}
    for _, header in headers:
        if "->" not in header:
            continue
        left, right = [p.strip() for p in header.split("->", 1)]
        if _is_hub(left) and not _is_hub(right):
            hub_sinks.setdefault(left, []).append(right)

    edges: list[tuple[str, str, int]] = []  # (left node, right node, source col)
    for col, header in headers:
        if "->" not in header:
            continue
        left, right = [p.strip() for p in header.split("->", 1)]
        if _is_hub(left):
            continue  # aggregate hub->sink edge; dropped
        if _is_hub(right):
            for sink in hub_sinks.get(right, []):
                edges.append((left, sink, col))
        else:
            edges.append((left, right, col))

    # (source col, output column name, direction). Sign follows the export
    # convention: the selected zone as start node -> positive ("from"), as end
    # node -> negated ("to"). Only flows to a non-selected neighbour are kept.
    # Neighbour naming: a country node ("XX_H2") becomes "XX00" in
    # "H2Exports_<main>_<CC>00"; every external source node (XDZ, XMA, XNO, XUA,
    # XAmmonia — codes starting with "X") is collapsed onto the main zone's single
    # "H2Exports_<main>" column, so all X.. flows sum (with the export sign).
    specs: list[tuple[int, str, str]] = []
    for left, right, col in edges:
        xc, yc = _cc(left), _cc(right)
        if xc in selected_countries and yc not in selected_countries:
            interested, neigh_raw, neigh_cc, direction = xc, right, yc, "from"
        elif yc in selected_countries and xc not in selected_countries:
            interested, neigh_raw, neigh_cc, direction = yc, left, xc, "to"
        else:
            continue
        main = main_zone_map.get(interested, f"{interested}00")
        if str(neigh_raw).startswith("X"):
            name = f"H2Exports_{main}_XX (MW/h)"        # aggregated external sources
        elif str(neigh_raw).endswith("_H2"):
            name = f"H2Exports_{main}_{neigh_cc}00 (MW/h)"
        else:
            name = f"H2Exports_{main}_{neigh_raw} (MW/h)"
        specs.append((col, name, direction))

    row_start = header_row + 1
    row_end = row_start + selected_hours
    col_data: dict[int, list] = {}
    for col in sorted({c for c, _, _ in specs}):
        col_data[col] = [
            v[0] for v in ws.iter_rows(
                min_row=row_start, max_row=row_end - 1,
                min_col=col, max_col=col, values_only=True,
            )
        ]
    wb.close()

    out: dict[str, list] = {}
    for col, name, direction in specs:
        raw = col_data[col]
        values = raw if direction == "from" else [(-v if v is not None else None) for v in raw]
        if name in out:  # sum flows mapping to the same column (X.. sources, hub sinks)
            out[name] = [(a or 0) + (b or 0) for a, b in zip(out[name], values)]
        else:
            out[name] = values

    return pd.DataFrame(out)


# ---------------------------------------------------------------------------
# Demand profile loaders
# ---------------------------------------------------------------------------


def load_electricity_demand_profiles(
    node_df: pd.DataFrame,
    selected_zones: list[str],
    scenario: int,
    climate_year: int,
    selected_hours: int,
) -> dict[str, list[dict]]:
    """Load hourly electricity demand profiles from ENTSO-E NT workbooks.

    Args:
        node_df (pd.DataFrame): Nodes table for zone iteration.
        selected_zones (list[str]): Zone codes to collect.
        scenario (int): Scenario year.  2050 is not supported and raises
            ``FileNotFoundError``.
        climate_year (int): Climate year to extract (e.g. ``2009``).
        selected_hours (int): Number of hourly values to read.

    Returns:
        dict[str, list[dict]]: Single-key dict
            ``{"Electricity Demand Profile": [{"Code": ..., "Year": ..., "Data": np.ndarray}]}``.

    Raises:
        FileNotFoundError: When *scenario* is ``2050``.

    Example:
        >>> profiles = load_electricity_demand_profiles(nodes, ["ES00"], 2030, 2009, 8736)
    """
    if scenario == 2050:
        raise FileNotFoundError("Electricity demand profiles are not available for scenario 2050")

    paths = {
        2030: r"inputs/Demand Profiles/NT/Electricity demand profiles/2030_National Trends.xlsx",
        2040: r"inputs/Demand Profiles/NT/Electricity demand profiles/2040_National Trends.xlsx",
    }
    sheet_year_row = 6
    results: list[dict] = []

    for code in node_df["Code"]:
        if code not in selected_zones:
            continue
        entry = _load_excel_demand_profile(
            filepath=paths[scenario],
            code=code,
            climate_year=climate_year,
            selected_hours=selected_hours,
            sheet_year_row=sheet_year_row,
            profile_key="Electricity Demand Profile",
        )
        results.append(entry)

    return {"Electricity Demand Profile": results}


def load_hydrogen_demand_profiles(
    node_df: pd.DataFrame,
    selected_zones: list[str],
    scenario: int,
    climate_year: int,
    selected_hours: int,
) -> dict[str, list[dict]]:
    """Load hourly hydrogen demand profiles from ENTSO-E NT workbooks.

    Args:
        node_df (pd.DataFrame): Nodes table for zone iteration.
        selected_zones (list[str]): Zone codes to collect.
        scenario (int): Scenario year.  2050 is not supported.
        climate_year (int): Climate year to extract.
        selected_hours (int): Number of hourly values to read.

    Returns:
        dict[str, list[dict]]: Single-key dict
            ``{"Hydrogen Demand Profile": [...]}``.

    Raises:
        FileNotFoundError: When *scenario* is ``2050``.

    Example:
        >>> profiles = load_hydrogen_demand_profiles(nodes, ["ES00"], 2030, 2009, 8736)
    """
    if scenario == 2050:
        raise FileNotFoundError("Hydrogen demand profiles are not available for scenario 2050")

    paths = {
        2030: r"inputs\Demand Profiles\NT\H2 demand profiles\H2 2030\NT_2030.xlsx",
        2040: r"inputs\Demand Profiles\NT\H2 demand profiles\H2 2040\NT_2040.xlsx",
    }
    sheet_year_row = 9
    results: list[dict] = []

    # Hydrogen has one demand sheet per country, named by that country's single H2
    # node — which may be any of the country's zone codes or "<CC>00" (e.g. LU00,
    # DKE1, SE01, ITN1). For a zone, try every same-country zone code plus "<CC>00"
    # and use the first candidate that is an actual sheet (one sheet per country).
    try:
        _h2_sheets = set(_excel(paths[scenario]).sheet_names)
    except Exception:
        _h2_sheets = set()
    country_zones: dict[str, list[str]] = {}
    for z in node_df["Code"]:
        country_zones.setdefault(str(z)[:2], []).append(str(z))

    def _h2_sheet_for(prefix: str) -> str | None:
        for cand in country_zones.get(prefix, []) + [f"{prefix}00"]:
            if cand in _h2_sheets:
                return cand
        return None

    _sheet_cache: dict[str, str | None] = {}
    for code in node_df["Code"]:
        if code not in selected_zones:
            continue
        pref = str(code)[:2]
        if pref not in _sheet_cache:
            _sheet_cache[pref] = _h2_sheet_for(pref)
        h2_sheet = _sheet_cache[pref]
        if h2_sheet is None:
            results.append({"Code": code, "Year": climate_year,
                            "Data": np.zeros(selected_hours)})
            continue
        entry = _load_excel_demand_profile(
            filepath=paths[scenario],
            code=h2_sheet,
            climate_year=climate_year,
            selected_hours=selected_hours,
            sheet_year_row=sheet_year_row,
            profile_key="Hydrogen Demand Profile",
        )
        entry["Code"] = code   # key by the electricity zone, not the H2 sheet name
        results.append(entry)

    return {"Hydrogen Demand Profile": results}


def load_gas_demand_profiles(
    node_df: pd.DataFrame,
    selected_zones: list[str],
    climate_year: int,
    selected_hours: int,
) -> dict[str, list[dict]]:
    """Return zero-filled gas demand profile entries for each selected zone.

    Gas hourly demand data is not yet available in the source dataset; zero
    arrays are recorded as placeholders to keep the output schema consistent.

    Args:
        node_df (pd.DataFrame): Nodes table for zone iteration.
        selected_zones (list[str]): Zone codes to include.
        climate_year (int): Climate year label stored in each entry.
        selected_hours (int): Length of the zero array.

    Returns:
        dict[str, list[dict]]: Single-key dict ``{"Gas Demand Profile": [...]}``.

    Example:
        >>> profiles = load_gas_demand_profiles(nodes, ["ES00"], 2009, 8736)
    """
    results: list[dict] = []
    for code in node_df["Code"]:
        if code not in selected_zones:
            continue
        results.append({"Code": code, "Year": climate_year, "Data": np.zeros(selected_hours)})
        print(f"Gas demand (zero placeholder) for {code}: recorded")
    return {"Gas Demand Profile": results}


def _load_excel_demand_profile(
    filepath: str,
    code: str,
    climate_year: int,
    selected_hours: int,
    sheet_year_row: int,
    profile_key: str,
) -> dict:
    """Read one zone's demand profile from an Excel workbook.

    Scans row *sheet_year_row* (0-indexed) for a column whose header equals
    *climate_year*, then extracts *selected_hours* values starting from the
    following row.
    """
    try:
        df = pd.read_excel(filepath, sheet_name=code, header=0)
    except (ValueError, FileNotFoundError):
        print(f"{profile_key} – worksheet '{code}' not found in {filepath}")
        return {"Code": code, "Year": 0, "Data": np.zeros(selected_hours)}

    if df.shape[0] <= sheet_year_row:
        return {"Code": code, "Year": climate_year, "Data": np.zeros(selected_hours)}

    year_row = df.iloc[sheet_year_row]
    col_idx = None
    for idx, val in zip(year_row.index[::-1], year_row.iloc[::-1]):
        if isinstance(val, (int, float)) and not pd.isna(val) and int(val) == climate_year:
            col_idx = idx
            break

    if col_idx is None:
        print(f"{profile_key} – climate year {climate_year} not found for {code}")
        return {"Code": code, "Year": climate_year, "Data": np.zeros(selected_hours)}

    data_start = sheet_year_row + 1
    profile_col = df.columns.get_loc(col_idx)
    raw = df.iloc[data_start: data_start + selected_hours, profile_col]
    arr = pd.to_numeric(raw, errors="coerce").to_numpy()
    if arr.shape[0] < selected_hours:
        arr = np.pad(arr, (0, selected_hours - arr.shape[0]), constant_values=np.nan)

    print(f"{profile_key} for {code}: OK")
    return {"Code": code, "Year": climate_year, "Data": arr}


# ---------------------------------------------------------------------------
# Generic PECD CSV profile loader
# ---------------------------------------------------------------------------


def _load_pecd_csv_profiles(
    node_df: pd.DataFrame,
    selected_zones: list[str],
    scenario: int,
    climate_year: int,
    selected_hours: int,
    profile_key: str,
    year_row_idx: int = 9,
    target_len: int | None = None,
) -> list[dict]:
    """Load hourly generation profiles from PECD CSV files for all selected zones.

    The PECD CSV files share a common layout: row *year_row_idx* (0-indexed)
    is a header row containing numeric climate year values; data starts on the
    following row.

    Args:
        node_df (pd.DataFrame): Nodes table for zone iteration.
        selected_zones (list[str]): Zone codes to collect.
        scenario (int): Scenario year used to resolve the file template.
        climate_year (int): Climate year column to extract.
        selected_hours (int): Number of hourly values to read (also used as
            the fallback zero-array length).
        profile_key (str): Key under which results are stored in the profile
            dict (e.g. ``'Solar Profile'``).
        year_row_idx (int): 0-indexed row number that contains climate year
            headers. Defaults to ``9``.
        target_len (int | None): Override the output array length; useful for
            Solar Rooftop which is always 8760 h. Defaults to *selected_hours*.

    Returns:
        list[dict]: List of ``{"Code": ..., "Year": ..., "Data": np.ndarray}``
            entries.

    Raises:
        KeyError: When *profile_key* is not in
            :data:`~collector.utils.config.PECD_FILE_TEMPLATES`.
    """
    if target_len is None:
        target_len = selected_hours

    file_template = PECD_FILE_TEMPLATES[profile_key][scenario]
    results: list[dict] = []

    for code in node_df["Code"]:
        if code not in selected_zones:
            continue
        path = file_template.format(code)
        try:
            df = pd.read_csv(path, header=0)
        except FileNotFoundError:
            print(f"{profile_key} for {code}: file not found – zeros used")
            results.append({"Code": code, "Year": 0, "Data": np.zeros(target_len)})
            continue

        col_idx = None
        if df.shape[0] > year_row_idx:
            year_row = df.iloc[year_row_idx]
            for idx, val in zip(year_row.index[::-1], year_row.iloc[::-1]):
                if isinstance(val, (int, float)) and not pd.isna(val) and int(val) == climate_year:
                    col_idx = idx
                    break

        if col_idx is None:
            print(f"{profile_key} for {code}: climate year {climate_year} not found – zeros used")
            results.append({"Code": code, "Year": climate_year, "Data": np.zeros(target_len)})
            continue

        data_start = year_row_idx + 1
        profile_col = df.columns.get_loc(col_idx)
        raw = df.iloc[data_start: data_start + target_len, profile_col]
        arr = pd.to_numeric(raw, errors="coerce").to_numpy()
        if arr.shape[0] < target_len:
            arr = np.pad(arr, (0, target_len - arr.shape[0]), constant_values=np.nan)

        results.append({"Code": code, "Year": climate_year, "Data": arr})
        print(f"{profile_key} for {code}: OK")

    return results


# ---------------------------------------------------------------------------
# Named PECD profile loaders (thin wrappers)
# ---------------------------------------------------------------------------


def load_csp_no_storage_profiles(
    node_df: pd.DataFrame, selected_zones: list[str],
    scenario: int, climate_year: int, selected_hours: int,
) -> dict[str, list[dict]]:
    """Load hourly CSP (no storage) generation profiles.

    Args:
        node_df (pd.DataFrame): Nodes table.
        selected_zones (list[str]): Zone codes.
        scenario (int): Scenario year.
        climate_year (int): Climate year.
        selected_hours (int): Number of hourly values.

    Returns:
        dict[str, list[dict]]: ``{"CSP_noStorage Profile": [...]}``.

    Example:
        >>> p = load_csp_no_storage_profiles(nodes, ["ES00"], 2030, 2009, 8736)
    """
    key = "CSP_noStorage Profile"
    return {key: _load_pecd_csv_profiles(node_df, selected_zones, scenario, climate_year, selected_hours, key)}


def load_csp_dispatch_profiles(
    node_df: pd.DataFrame, selected_zones: list[str],
    scenario: int, climate_year: int, selected_hours: int,
) -> dict[str, list[dict]]:
    """Load hourly CSP (with storage – dispatch) generation profiles.

    Args:
        node_df (pd.DataFrame): Nodes table.
        selected_zones (list[str]): Zone codes.
        scenario (int): Scenario year.
        climate_year (int): Climate year.
        selected_hours (int): Number of hourly values.

    Returns:
        dict[str, list[dict]]: ``{"CSP_withStorage_D Profile": [...]}``.

    Example:
        >>> p = load_csp_dispatch_profiles(nodes, ["ES00"], 2030, 2009, 8736)
    """
    key = "CSP_withStorage_D Profile"
    return {key: _load_pecd_csv_profiles(node_df, selected_zones, scenario, climate_year, selected_hours, key)}


def load_csp_predispatch_profiles(
    node_df: pd.DataFrame, selected_zones: list[str],
    scenario: int, climate_year: int, selected_hours: int,
) -> dict[str, list[dict]]:
    """Load hourly CSP (with storage – pre-dispatch) generation profiles.

    Args:
        node_df (pd.DataFrame): Nodes table.
        selected_zones (list[str]): Zone codes.
        scenario (int): Scenario year.
        climate_year (int): Climate year.
        selected_hours (int): Number of hourly values.

    Returns:
        dict[str, list[dict]]: ``{"CSP_withStorage_PreD Profile": [...]}``.

    Example:
        >>> p = load_csp_predispatch_profiles(nodes, ["ES00"], 2030, 2009, 8736)
    """
    key = "CSP_withStorage_PreD Profile"
    return {key: _load_pecd_csv_profiles(node_df, selected_zones, scenario, climate_year, selected_hours, key)}


def load_solar_profiles(
    node_df: pd.DataFrame, selected_zones: list[str],
    scenario: int, climate_year: int, selected_hours: int,
) -> dict[str, list[dict]]:
    """Load hourly utility-scale solar PV capacity factor profiles.

    Args:
        node_df (pd.DataFrame): Nodes table.
        selected_zones (list[str]): Zone codes.
        scenario (int): Scenario year.
        climate_year (int): Climate year.
        selected_hours (int): Number of hourly values.

    Returns:
        dict[str, list[dict]]: ``{"Solar Profile": [...]}``.

    Example:
        >>> p = load_solar_profiles(nodes, ["ES00"], 2030, 2009, 8736)
    """
    key = "Solar Profile"
    return {key: _load_pecd_csv_profiles(node_df, selected_zones, scenario, climate_year, selected_hours, key)}


def load_solar_rooftop_profiles(
    node_df: pd.DataFrame, selected_zones: list[str],
    scenario: int, climate_year: int, selected_hours: int,
) -> dict[str, list[dict]]:
    """Load hourly rooftop solar PV capacity factor profiles.

    Rooftop PV files always contain 8760 values; the array is padded or
    truncated to that length rather than *selected_hours*.

    Args:
        node_df (pd.DataFrame): Nodes table.
        selected_zones (list[str]): Zone codes.
        scenario (int): Scenario year.
        climate_year (int): Climate year.
        selected_hours (int): Passed for API consistency (not used for length).

    Returns:
        dict[str, list[dict]]: ``{"Solar_Rooftop Profile": [...]}``.

    Example:
        >>> p = load_solar_rooftop_profiles(nodes, ["ES00"], 2030, 2009, 8736)
    """
    key = "Solar_Rooftop Profile"
    return {key: _load_pecd_csv_profiles(
        node_df, selected_zones, scenario, climate_year, selected_hours, key,
        target_len=SOLAR_ROOFTOP_TARGET_LEN,
    )}


def load_solar_utility_profiles(
    node_df: pd.DataFrame, selected_zones: list[str],
    scenario: int, climate_year: int, selected_hours: int,
) -> dict[str, list[dict]]:
    """Load hourly utility-scale solar PV (utility sub-type) profiles.

    Args:
        node_df (pd.DataFrame): Nodes table.
        selected_zones (list[str]): Zone codes.
        scenario (int): Scenario year.
        climate_year (int): Climate year.
        selected_hours (int): Number of hourly values.

    Returns:
        dict[str, list[dict]]: ``{"Solar_Utility Profile": [...]}``.

    Example:
        >>> p = load_solar_utility_profiles(nodes, ["ES00"], 2030, 2009, 8736)
    """
    key = "Solar_Utility Profile"
    return {key: _load_pecd_csv_profiles(node_df, selected_zones, scenario, climate_year, selected_hours, key)}


def load_wind_offshore_profiles(
    node_df: pd.DataFrame, selected_zones: list[str],
    scenario: int, climate_year: int, selected_hours: int,
) -> dict[str, list[dict]]:
    """Load hourly offshore wind capacity factor profiles.

    Args:
        node_df (pd.DataFrame): Nodes table.
        selected_zones (list[str]): Zone codes.
        scenario (int): Scenario year.
        climate_year (int): Climate year.
        selected_hours (int): Number of hourly values.

    Returns:
        dict[str, list[dict]]: ``{"Wind_Offshore Profile": [...]}``.

    Example:
        >>> p = load_wind_offshore_profiles(nodes, ["ES00"], 2030, 2009, 8736)
    """
    key = "Wind_Offshore Profile"
    return {key: _load_pecd_csv_profiles(node_df, selected_zones, scenario, climate_year, selected_hours, key)}


def load_wind_onshore_profiles(
    node_df: pd.DataFrame, selected_zones: list[str],
    scenario: int, climate_year: int, selected_hours: int,
) -> dict[str, list[dict]]:
    """Load hourly onshore wind capacity factor profiles.

    Args:
        node_df (pd.DataFrame): Nodes table.
        selected_zones (list[str]): Zone codes.
        scenario (int): Scenario year.
        climate_year (int): Climate year.
        selected_hours (int): Number of hourly values.

    Returns:
        dict[str, list[dict]]: ``{"Wind_Onshore Profile": [...]}``.

    Example:
        >>> p = load_wind_onshore_profiles(nodes, ["ES00"], 2030, 2009, 8736)
    """
    key = "Wind_Onshore Profile"
    return {key: _load_pecd_csv_profiles(node_df, selected_zones, scenario, climate_year, selected_hours, key)}


# ---------------------------------------------------------------------------
# Generic hydro inflow loader
# ---------------------------------------------------------------------------


def _load_hydro_inflow_profiles(
    node_df: pd.DataFrame,
    selected_zones: list[str],
    scenario: int,
    climate_year: int,
    profile_key: str,
    scale_factor: float = HYDRO_SCALE_FACTOR,
) -> list[dict]:
    """Load daily or weekly hydro inflow energy series from PEMMDB files.

    Looks up the sheet name and expected series length from the config
    constants :data:`~collector.utils.config.HYDRO_SHEET_NAMES` and
    :data:`~collector.utils.config.HYDRO_TARGET_LENGTHS`.

    Args:
        node_df (pd.DataFrame): Nodes table for zone iteration.
        selected_zones (list[str]): Zone codes to collect.
        scenario (int): Scenario year.
        climate_year (int): Climate year column to extract.
        profile_key (str): Key identifying the hydro type (e.g.
            ``'River Flow Energy'``).
        scale_factor (float): Multiplier applied to raw values (converts
            GWh → MWh by default). Defaults to ``1000.0``.

    Returns:
        list[dict]: List of ``{"Code": ..., "Year": ..., "Data": np.ndarray}``
            entries.
    """
    file_template = HYDRO_FILE_TEMPLATES[scenario]
    sheet_name = HYDRO_SHEET_NAMES[profile_key]
    target_len = HYDRO_TARGET_LENGTHS[profile_key]
    results: list[dict] = []

    for code in node_df["Code"]:
        if code not in selected_zones:
            continue
        path = file_template.format(code)
        try:
            df = pd.read_excel(_excel(path), header=0, sheet_name=sheet_name)
        except FileNotFoundError:
            print(f"{profile_key} for {code}: file not found – zeros used")
            results.append({"Code": code, "Year": 0, "Data": np.zeros(target_len)})
            continue

        year_row = df.iloc[0]
        col_idx = None
        for idx, val in zip(year_row.index[::-1], year_row.iloc[::-1]):
            if isinstance(val, (int, float)) and not pd.isna(val) and int(val) == climate_year:
                col_idx = idx
                break

        if col_idx is None:
            print(f"{profile_key} for {code}: climate year {climate_year} not found – zeros used")
            results.append({"Code": code, "Year": climate_year, "Data": np.zeros(target_len)})
            continue

        profile_col = df.columns.get_loc(col_idx)
        raw = df.iloc[1: 1 + target_len, profile_col]
        arr = pd.to_numeric(raw, errors="coerce").to_numpy()
        if arr.shape[0] < target_len:
            arr = np.pad(arr, (0, target_len - arr.shape[0]), constant_values=np.nan)

        results.append({"Code": code, "Year": climate_year, "Data": arr * scale_factor})
        print(f"{profile_key} for {code}: OK")

    return results


# ---------------------------------------------------------------------------
# Named hydro inflow loaders (thin wrappers)
# ---------------------------------------------------------------------------


def load_river_flow_profiles(
    node_df: pd.DataFrame, selected_zones: list[str],
    scenario: int, climate_year: int,
) -> dict[str, list[dict]]:
    """Load daily run-of-river energy inflow profiles.

    Args:
        node_df (pd.DataFrame): Nodes table.
        selected_zones (list[str]): Zone codes.
        scenario (int): Scenario year.
        climate_year (int): Climate year.

    Returns:
        dict[str, list[dict]]: ``{"River Flow Energy": [...]}``.

    Example:
        >>> p = load_river_flow_profiles(nodes, ["ES00"], 2030, 2009)
    """
    key = "River Flow Energy"
    return {key: _load_hydro_inflow_profiles(node_df, selected_zones, scenario, climate_year, key)}


def load_pondage_flow_profiles(
    node_df: pd.DataFrame, selected_zones: list[str],
    scenario: int, climate_year: int,
) -> dict[str, list[dict]]:
    """Load daily pondage hydro energy inflow profiles.

    Args:
        node_df (pd.DataFrame): Nodes table.
        selected_zones (list[str]): Zone codes.
        scenario (int): Scenario year.
        climate_year (int): Climate year.

    Returns:
        dict[str, list[dict]]: ``{"Pondage Flow Energy": [...]}``.

    Example:
        >>> p = load_pondage_flow_profiles(nodes, ["ES00"], 2030, 2009)
    """
    key = "Pondage Flow Energy"
    return {key: _load_hydro_inflow_profiles(node_df, selected_zones, scenario, climate_year, key)}


def load_reservoir_flow_profiles(
    node_df: pd.DataFrame, selected_zones: list[str],
    scenario: int, climate_year: int,
) -> dict[str, list[dict]]:
    """Load weekly reservoir hydro energy inflow profiles.

    Args:
        node_df (pd.DataFrame): Nodes table.
        selected_zones (list[str]): Zone codes.
        scenario (int): Scenario year.
        climate_year (int): Climate year.

    Returns:
        dict[str, list[dict]]: ``{"Reservoir Flow Energy": [...]}``.

    Example:
        >>> p = load_reservoir_flow_profiles(nodes, ["ES00"], 2030, 2009)
    """
    key = "Reservoir Flow Energy"
    return {key: _load_hydro_inflow_profiles(node_df, selected_zones, scenario, climate_year, key)}


def load_open_ps_flow_profiles(
    node_df: pd.DataFrame, selected_zones: list[str],
    scenario: int, climate_year: int,
) -> dict[str, list[dict]]:
    """Load weekly open-loop pump-storage energy inflow profiles.

    Args:
        node_df (pd.DataFrame): Nodes table.
        selected_zones (list[str]): Zone codes.
        scenario (int): Scenario year.
        climate_year (int): Climate year.

    Returns:
        dict[str, list[dict]]: ``{"Open_PS Flow Energy": [...]}``.

    Example:
        >>> p = load_open_ps_flow_profiles(nodes, ["ES00"], 2030, 2009)
    """
    key = "Open_PS Flow Energy"
    return {key: _load_hydro_inflow_profiles(node_df, selected_zones, scenario, climate_year, key)}


def load_closed_ps_flow_profiles(
    node_df: pd.DataFrame, selected_zones: list[str],
    scenario: int, climate_year: int,
) -> dict[str, list[dict]]:
    """Load weekly closed-loop pump-storage energy inflow profiles.

    Args:
        node_df (pd.DataFrame): Nodes table.
        selected_zones (list[str]): Zone codes.
        scenario (int): Scenario year.
        climate_year (int): Climate year.

    Returns:
        dict[str, list[dict]]: ``{"Closed_PS Flow Energy": [...]}``.

    Example:
        >>> p = load_closed_ps_flow_profiles(nodes, ["ES00"], 2030, 2009)
    """
    key = "Closed_PS Flow Energy"
    return {key: _load_hydro_inflow_profiles(node_df, selected_zones, scenario, climate_year, key)}


# ---------------------------------------------------------------------------
# Convenience: load everything at once
# ---------------------------------------------------------------------------


def load_all_profiles(
    node_df: pd.DataFrame,
    selected_zones: list[str],
    scenario: int,
    climate_year: int,
    selected_hours: int,
) -> dict[str, list[dict]]:
    """Load all generation and demand profiles for the selected zones.

    Calls every individual profile loader and merges their results into a
    single dict keyed by profile type.  Missing files produce zero-filled
    entries so downstream code always receives a complete dict.

    Args:
        node_df (pd.DataFrame): Nodes table.
        selected_zones (list[str]): Zone codes to collect.
        scenario (int): Scenario year.
        climate_year (int): Climate year.
        selected_hours (int): Number of hourly values per time series.

    Returns:
        dict[str, list[dict]]: Combined profiles dict with one key per profile
            type and one list entry per zone.

    Example:
        >>> nodes = load_nodes()
        >>> all_profiles = load_all_profiles(nodes, ["ES00"], 2030, 2009, 8736)
        >>> list(all_profiles.keys())[:3]
        ['Electricity Demand Profile', 'Hydrogen Demand Profile', 'Gas Demand Profile']
    """
    combined: dict[str, list[dict]] = {}
    loaders = [
        lambda: load_electricity_demand_profiles(node_df, selected_zones, scenario, climate_year, selected_hours),
        lambda: load_hydrogen_demand_profiles(node_df, selected_zones, scenario, climate_year, selected_hours),
        lambda: load_gas_demand_profiles(node_df, selected_zones, climate_year, selected_hours),
        lambda: load_csp_no_storage_profiles(node_df, selected_zones, scenario, climate_year, selected_hours),
        lambda: load_csp_dispatch_profiles(node_df, selected_zones, scenario, climate_year, selected_hours),
        lambda: load_csp_predispatch_profiles(node_df, selected_zones, scenario, climate_year, selected_hours),
        lambda: load_solar_profiles(node_df, selected_zones, scenario, climate_year, selected_hours),
        lambda: load_solar_rooftop_profiles(node_df, selected_zones, scenario, climate_year, selected_hours),
        lambda: load_solar_utility_profiles(node_df, selected_zones, scenario, climate_year, selected_hours),
        lambda: load_wind_offshore_profiles(node_df, selected_zones, scenario, climate_year, selected_hours),
        lambda: load_wind_onshore_profiles(node_df, selected_zones, scenario, climate_year, selected_hours),
        lambda: load_river_flow_profiles(node_df, selected_zones, scenario, climate_year),
        lambda: load_pondage_flow_profiles(node_df, selected_zones, scenario, climate_year),
        lambda: load_reservoir_flow_profiles(node_df, selected_zones, scenario, climate_year),
        lambda: load_open_ps_flow_profiles(node_df, selected_zones, scenario, climate_year),
        lambda: load_closed_ps_flow_profiles(node_df, selected_zones, scenario, climate_year),
    ]

    for loader in loaders:
        try:
            combined.update(loader())
        except (FileNotFoundError, KeyError) as exc:
            print(f"Warning: profile loader skipped – {exc}")

    return combined


def load_commodity_prices(
    scenario_year: int,
    filepath: str = FILEPATH_COMMODITY_PRICES,
) -> dict[str, float]:
    """Load TYNDP 2024 fuel commodity prices from the Matrix 2024 sheet.

    Returns a dict of fuel-key → price in EUR/MWh (converted from EUR/GJ ×3.6).
    Returns an empty dict if the file cannot be read.

    Args:
        scenario_year (int): Scenario year (2030, 2040, or 2050).
        filepath (str): Path to the TYNDP commodity prices workbook.

    Returns:
        dict[str, float]: Fuel prices in EUR/MWh keyed by fuel name.

    Example:
        >>> prices = load_commodity_prices(2030)
        >>> prices["Natural_Gas"]
        22.64
    """
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as exc:
        print(f"Warning: commodity prices file not found – {exc}")
        return {}

    ws = wb["Matrix 2024"]
    year_col: int | None = None
    fuel_prices: dict[str, float] = {}

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 2:
            for j, val in enumerate(row):
                if val == scenario_year:
                    year_col = j
                    break
            if year_col is None:
                year_col = 3  # default to 2030 column
        elif i >= 3 and year_col is not None:
            fuel = row[1]
            if fuel is None:
                continue
            price = row[year_col]
            if price is not None:
                try:
                    fuel_prices[str(fuel).strip()] = float(price) * 3.6
                except (TypeError, ValueError):
                    pass
        if i > 25:
            break

    def _p(key: str) -> float:
        return round(fuel_prices.get(key, 0.0), 4)

    return {
        "Nuclear":     _p("Nuclear"),
        "Hard_coal":   _p("Hard coal"),
        "Lignite_G1":  _p("Lignite G1 (BG - MK - CZ)"),
        "Lignite_G2":  _p("Lignite G2 (SK - DE - RS - PL - ME - UKNI - BA - IE)"),
        "Lignite_G3":  _p("Lignite G3 (SL - RO - HU)"),
        "Lignite_G4":  _p("Lignite G4 (GR - TR)"),
        "Natural_Gas": _p("Natural Gas"),
        "Crude_oil":   _p("Crude oil"),
        "Light_oil":   _p("Light oil"),
        "Heavy_oil":   _p("Heavy oil"),
        "Oil_shale":   _p("Oil Shale"),
        "Hydrogen":    _p("Hydrogen (blue )"),
        "Biomethane":  _p("Biomethane"),
        "Gas_blend_NT": _p("Gas (blend of biomethane, synthetic gas and NG) NT+"),
        # CO2 price is quoted in EUR/ton (not EUR/GJ) → undo the ×3.6 conversion
        "CO2_price":   round(fuel_prices.get("CO2 price", 0.0) / 3.6, 4),
    }


def load_lignite_groups(
    filepath: str = FILEPATH_COMMODITY_PRICES,
) -> dict[str, str]:
    """Parse Lignite country-to-group mapping from the Matrix 2024 sheet.

    Reads the fuel name strings (e.g. "Lignite G2 (SK - DE - RS - PL - ME - UKNI - BA - IE)")
    and extracts each country code from the parentheses.  Returns a dict mapping
    country code → price key (e.g. ``{"DE": "Lignite_G2", "UKNI": "Lignite_G2", ...}``).

    Using the Excel as the source means new groups or countries added to the
    workbook are picked up automatically without any code change.

    Args:
        filepath (str): Path to the TYNDP commodity prices workbook.

    Returns:
        dict[str, str]: Country code → lignite price key.

    Example:
        >>> groups = load_lignite_groups()
        >>> groups["UKNI"]
        'Lignite_G2'
    """
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as exc:
        print(f"Warning: commodity prices file not found – {exc}")
        return {}

    ws = wb["Matrix 2024"]
    result: dict[str, str] = {}

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3:
            continue
        fuel = str(row[1]).strip() if row[1] else ""
        m = re.match(r"Lignite\s+G(\d+)\s*\(([^)]+)\)", fuel)
        if m:
            key = f"Lignite_G{m.group(1)}"
            for country in re.split(r"\s*-\s*", m.group(2)):
                country = country.strip()
                if country:
                    result[country] = key
        if i > 10:
            break

    return result
